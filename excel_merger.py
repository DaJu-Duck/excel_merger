# -*- coding: utf-8 -*-
import sys
import os
import subprocess
import importlib
import time
import shutil  # 添加导入shutil模块
import copy  # 添加导入copy模块
from threading import Thread
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.translate import Translator
from functools import partial

# 尝试导入PyQt，如果失败，安装基本的依赖
try:
    from PyQt5.QtWidgets import (QApplication, QDialog, QProgressBar, QLabel, 
                              QPushButton, QVBoxLayout, QHBoxLayout, QTextEdit, 
                              QMainWindow, QFileDialog, QGroupBox, QScrollArea, QWidget,
                              QListWidget, QListWidgetItem, QCheckBox, QComboBox, 
                              QMessageBox, QLineEdit, QSizePolicy, QGridLayout, QTabWidget,
                              QFormLayout, QRadioButton, QTableWidget, QTableWidgetItem,
                              QHeaderView, QDialogButtonBox)
    from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer, QSize
    from PyQt5.QtGui import QFont
    HAS_PYQT = True
except ImportError:
    HAS_PYQT = False
    print("PyQt5 未安装，将尝试安装...")

def install_package(package):
    """安装指定的包"""
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        return True
    except subprocess.CalledProcessError:
        return False

def is_package_installed(package_name):
    """检查包是否已安装"""
    try:
        importlib.import_module(package_name)
        return True
    except ImportError:
        return False

# 需要检查的依赖列表
REQUIRED_PACKAGES = [
    {'name': 'PyQt5', 'import_name': 'PyQt5', 'install_name': 'PyQt5'},
    {'name': 'pandas', 'import_name': 'pandas', 'install_name': 'pandas'},
    {'name': 'openpyxl', 'import_name': 'openpyxl', 'install_name': 'openpyxl'},
]

# 检测缺失的依赖
MISSING_PACKAGES = []
for package in REQUIRED_PACKAGES:
    if not is_package_installed(package['import_name']):
        MISSING_PACKAGES.append(package)

# 如果PyQt5未安装，需要先安装它才能显示UI
if not HAS_PYQT:
    print("正在安装PyQt5...")
    if install_package('PyQt5'):
        print("PyQt5安装成功，请重启应用")
    else:
        print("PyQt5安装失败，请手动安装后再运行: pip install PyQt5")
    sys.exit(1)

# 现在可以安全地导入PyQt相关的模块
from PyQt5.QtWidgets import (QApplication, QDialog, QProgressBar, QLabel, 
                          QPushButton, QVBoxLayout, QHBoxLayout, QTextEdit)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QTimer

class DependencyInstaller(QThread):
    """后台线程用于安装依赖包"""
    progress_signal = pyqtSignal(str, int)
    finished_signal = pyqtSignal(bool)
    
    def __init__(self, packages):
        super().__init__()
        self.packages = packages
        
    def run(self):
        """运行安装过程"""
        total = len(self.packages)
        success = True
        
        for i, package in enumerate(self.packages):
            progress = int((i / total) * 100)
            self.progress_signal.emit(f"正在安装 {package['name']}...", progress)
            
            if not install_package(package['install_name']):
                self.progress_signal.emit(f"安装 {package['name']} 失败", progress)
                success = False
                break
            
            self.progress_signal.emit(f"{package['name']} 安装成功", min(progress + 10, 99))
        
        if success:
            self.progress_signal.emit("所有依赖项安装完成", 100)
        
        self.finished_signal.emit(success)

class DependencyDialog(QDialog):
    """依赖安装对话框"""
    def __init__(self, missing_packages):
        super().__init__()
        self.missing_packages = missing_packages
        self.installer = None
        self.initUI()
        
    def initUI(self):
        """初始化对话框界面"""
        self.setWindowTitle("依赖项安装")
        self.setMinimumWidth(500)
        self.setMinimumHeight(300)
        
        layout = QVBoxLayout()
        
        # 头部信息
        top_label = QLabel("需要安装以下依赖项才能运行此应用：")
        layout.addWidget(top_label)
        
        # 缺失依赖列表
        self.text_area = QTextEdit()
        self.text_area.setReadOnly(True)
        package_text = "\n".join([f"• {p['name']}" for p in self.missing_packages])
        self.text_area.setText(package_text)
        layout.addWidget(self.text_area)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setValue(0)
        layout.addWidget(self.progress)
        
        # 状态标签
        self.status_label = QLabel("准备安装...")
        layout.addWidget(self.status_label)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        self.install_btn = QPushButton("安装依赖项")
        self.install_btn.clicked.connect(self.install_dependencies)
        btn_layout.addWidget(self.install_btn)
        
        self.manual_btn = QPushButton("手动安装说明")
        self.manual_btn.clicked.connect(self.show_manual_instructions)
        btn_layout.addWidget(self.manual_btn)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def install_dependencies(self):
        """开始安装依赖项"""
        self.install_btn.setEnabled(False)
        self.manual_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        
        # 在后台线程中安装依赖
        self.installer = DependencyInstaller(self.missing_packages)
        self.installer.progress_signal.connect(self.update_progress)
        self.installer.finished_signal.connect(self.installation_finished)
        self.installer.start()
    
    def update_progress(self, message, value):
        """更新安装进度"""
        self.status_label.setText(message)
        self.progress.setValue(value)
    
    def installation_finished(self, success):
        """安装完成处理"""
        if success:
            self.status_label.setText("安装成功！应用将在3秒后重启...")
            QTimer.singleShot(3000, lambda: self.accept())
        else:
            self.status_label.setText("安装失败。请尝试手动安装。")
            self.manual_btn.setEnabled(True)
            self.cancel_btn.setEnabled(True)
    
    def show_manual_instructions(self):
        """显示手动安装说明"""
        instructions = "请在命令行中运行以下命令安装依赖：\n\n"
        for package in self.missing_packages:
            instructions += f"pip install {package['install_name']}\n"
        
        self.text_area.setText(instructions)
    
    def closeEvent(self, event):
        """关闭事件处理"""
        if self.installer and self.installer.isRunning():
            self.installer.terminate()
        event.accept()

def check_dependencies():
    """检查依赖并启动安装过程"""
    # 使用前面已经检测到的缺失依赖
    if MISSING_PACKAGES:
        # 创建应用实例以显示对话框
        app = QApplication.instance()
        if app is None:
            app = QApplication(sys.argv)
            created_app = True
        else:
            created_app = False
            
        dialog = DependencyDialog(MISSING_PACKAGES)
        
        # 在演示模式下修改对话框
        if DEMO_MODE:
            dialog.text_area.setText("演示模式：这是依赖安装对话框的演示，实际上不会安装任何依赖")
            dialog.status_label.setText("演示模式 - 所有依赖已实际安装，这只是界面演示")
            # 修改安装按钮的行为
            dialog.install_btn.clicked.disconnect()
            dialog.install_btn.clicked.connect(lambda: dialog.update_progress("这只是演示，没有实际安装依赖", 50))
        
        result = dialog.exec_()
        
        if not DEMO_MODE and result == QDialog.Accepted:
            # 安装成功，重启应用
            print("依赖安装成功，重启应用...")
            if created_app:
                app.quit()
            # 使用os.execv来重启程序
            os.execv(sys.executable, [sys.executable] + sys.argv)
            return True
        else:
            # 用户取消了安装或者是演示模式
            if DEMO_MODE:
                return True  # 演示模式继续运行程序
            print("依赖安装被取消")
            return False
    else:
        # 没有缺失的依赖
        return True

# 主程序入口前检查依赖
if not check_dependencies():
    sys.exit(1)

# 导入已安装的依赖
import pandas as pd
from functools import partial

class RelationDialog(QDialog):
    """用于处理关联设置的对话框"""
    def __init__(self, relation_type, file_data, parent=None):
        super().__init__(parent)
        self.relation_type = relation_type
        self.file_data = file_data
        self.result_data = {}
        
        # 设置对话框基本属性
        if relation_type == "single":
            self.setWindowTitle("单一关联设置")
        elif relation_type == "chain":
            self.setWindowTitle("链式关联设置")
        else:
            self.setWindowTitle("星形关联设置")
            
        self.setMinimumSize(700, 500)
        
        # 创建主布局
        main_layout = QVBoxLayout(self)
        
        # 添加说明标签
        description = QLabel()
        if relation_type == "single":
            description.setText("设置每个表格对应的关联字段。单一关联允许每个表格使用不同名称的字段进行关联。")
        elif relation_type == "chain":
            description.setText("设置表格之间的链接关系。您可以指定哪些表格相互关联及其关联字段。")
        else:
            description.setText("设置中心表与其他表的关联。您可以指定哪个表为中心表，并设置与其他表的关联字段。")
        
        description.setWordWrap(True)
        main_layout.addWidget(description)
        
        # 根据关联类型创建不同的设置界面
        if relation_type == "single":
            self.setup_single_relation(main_layout)
        elif relation_type == "chain":
            self.setup_chain_relation(main_layout)
        else:
            self.setup_star_relation(main_layout)
            
        # 添加按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        main_layout.addWidget(buttons)
        
    def setup_single_relation(self, layout):
        """设置单一关联的界面"""
        # 使用表格来显示文件与关联字段的映射
        table = QTableWidget(len(self.file_data), 2)
        table.setHorizontalHeaderLabels(["文件名", "关联字段"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        
        # 初始化关联字段数据
        self.single_field_combos = []
        
        # 填充表格
        for i, file_info in enumerate(self.file_data):
            # 文件名
            name_item = QTableWidgetItem(file_info['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)  # 不可编辑
            table.setItem(i, 0, name_item)
            
            # 关联字段下拉框
            combo = QComboBox()
            combo.addItems(file_info['columns'])
            combo.setProperty("file_index", i)
            self.single_field_combos.append(combo)
            
            # 如果之前已有设置，则选中对应字段
            if 'ref_field' in file_info:
                idx = combo.findText(file_info['ref_field'])
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            
            table.setCellWidget(i, 1, combo)
        
        layout.addWidget(table)
        
        # 添加字段匹配建议
        sugg_layout = QHBoxLayout()
        sugg_label = QLabel("字段匹配建议:")
        sugg_layout.addWidget(sugg_label)
        
        self.suggest_combo = QComboBox()
        common_fields = ["工号", "ID", "编号", "人员编号", "员工编号", "学号"]
        self.suggest_combo.addItems(common_fields)
        self.suggest_combo.currentTextChanged.connect(self.apply_suggestion)
        sugg_layout.addWidget(self.suggest_combo)
        
        apply_btn = QPushButton("应用建议")
        apply_btn.clicked.connect(lambda: self.apply_suggestion(self.suggest_combo.currentText()))
        sugg_layout.addWidget(apply_btn)
        
        layout.addLayout(sugg_layout)
        
    def apply_suggestion(self, suggestion):
        """根据建议自动匹配字段"""
        for i, combo in enumerate(self.single_field_combos):
            # 获取当前文件的列
            file_info = self.file_data[i]
            columns = file_info['columns']
            
            # 尝试找到与建议匹配的列
            for col in columns:
                if suggestion.lower() in col.lower():
                    combo.setCurrentText(col)
                    break
    
    def setup_chain_relation(self, layout):
        """设置链式关联的界面"""
        # 文件列表
        files_group = QGroupBox("文件列表")
        files_layout = QVBoxLayout(files_group)
        
        self.files_list = QListWidget()
        for i, file_info in enumerate(self.file_data):
            self.files_list.addItem(f"{i+1}: {file_info['name']}")
        files_layout.addWidget(self.files_list)
        
        layout.addWidget(files_group)
        
        # 关联关系设置
        relations_group = QGroupBox("关联关系")
        relations_layout = QVBoxLayout(relations_group)
        
        # 关联控制
        control_layout = QHBoxLayout()
        
        # 选择源文件和目标文件
        source_layout = QHBoxLayout()
        source_layout.addWidget(QLabel("源文件:"))
        self.source_combo = QComboBox()
        for i, file_info in enumerate(self.file_data):
            self.source_combo.addItem(file_info['name'], i)  # 存储索引作为用户数据
        source_layout.addWidget(self.source_combo, 1)
        control_layout.addLayout(source_layout)
        
        target_layout = QHBoxLayout()
        target_layout.addWidget(QLabel("目标文件:"))
        self.target_combo = QComboBox()
        for i, file_info in enumerate(self.file_data):
            self.target_combo.addItem(file_info['name'], i)
        if len(self.file_data) > 1:
            self.target_combo.setCurrentIndex(1)  # 默认选择第二个文件
        target_layout.addWidget(self.target_combo, 1)
        control_layout.addLayout(target_layout)
        
        relations_layout.addLayout(control_layout)
        
        # 字段选择
        fields_layout = QHBoxLayout()
        
        source_field_layout = QHBoxLayout()
        source_field_layout.addWidget(QLabel("源字段:"))
        self.source_field_combo = QComboBox()
        if len(self.file_data) > 0:
            self.source_field_combo.addItems(self.file_data[0]['columns'])
        source_field_layout.addWidget(self.source_field_combo, 1)
        fields_layout.addLayout(source_field_layout)
        
        target_field_layout = QHBoxLayout()
        target_field_layout.addWidget(QLabel("目标字段:"))
        self.target_field_combo = QComboBox()
        if len(self.file_data) > 1:
            self.target_field_combo.addItems(self.file_data[1]['columns'])
        target_field_layout.addWidget(self.target_field_combo, 1)
        fields_layout.addLayout(target_field_layout)
        
        relations_layout.addLayout(fields_layout)
        
        # 添加与删除关系的按钮
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("添加关联")
        add_btn.clicked.connect(self.add_chain_relation)
        btn_layout.addWidget(add_btn)
        
        remove_btn = QPushButton("删除关联")
        remove_btn.clicked.connect(self.remove_chain_relation)
        btn_layout.addWidget(remove_btn)
        
        relations_layout.addLayout(btn_layout)
        
        # 已添加的关联关系列表
        self.relations_list = QListWidget()
        relations_layout.addWidget(self.relations_list)
        
        layout.addWidget(relations_group)
        
        # 连接源文件和目标文件的变化事件
        self.source_combo.currentIndexChanged.connect(self.update_source_fields)
        self.target_combo.currentIndexChanged.connect(self.update_target_fields)
        
        # 初始化链式关联数据结构
        self.chain_relations = []
        
    def update_source_fields(self, index):
        """更新源文件的字段列表"""
        file_idx = self.source_combo.currentData()
        if file_idx is not None and 0 <= file_idx < len(self.file_data):
            self.source_field_combo.clear()
            self.source_field_combo.addItems(self.file_data[file_idx]['columns'])
    
    def update_target_fields(self, index):
        """更新目标文件的字段列表"""
        file_idx = self.target_combo.currentData()
        if file_idx is not None and 0 <= file_idx < len(self.file_data):
            self.target_field_combo.clear()
            self.target_field_combo.addItems(self.file_data[file_idx]['columns'])
    
    def add_chain_relation(self):
        """添加链式关联关系"""
        source_idx = self.source_combo.currentData()
        target_idx = self.target_combo.currentData()
        
        # 检查是否选择了相同的文件
        if source_idx == target_idx:
            QMessageBox.warning(self, "警告", "不能关联相同的文件")
            return
            
        source_name = self.file_data[source_idx]['name']
        target_name = self.file_data[target_idx]['name']
        source_field = self.source_field_combo.currentText()
        target_field = self.target_field_combo.currentText()
        
        # 检查是否已存在相同的关联
        for relation in self.chain_relations:
            if (relation['source_idx'] == source_idx and 
                relation['target_idx'] == target_idx and
                relation['source_field'] == source_field and
                relation['target_field'] == target_field):
                QMessageBox.information(self, "提示", "该关联关系已存在")
                return
        
        # 添加到关联列表
        relation = {
            'source_idx': source_idx,
            'target_idx': target_idx,
            'source_field': source_field,
            'target_field': target_field
        }
        self.chain_relations.append(relation)
        
        # 更新界面
        relation_text = f"{source_name}[{source_field}] → {target_name}[{target_field}]"
        self.relations_list.addItem(relation_text)
    
    def remove_chain_relation(self):
        """删除选中的链式关联关系"""
        current_row = self.relations_list.currentRow()
        if current_row >= 0 and current_row < len(self.chain_relations):
            del self.chain_relations[current_row]
            self.relations_list.takeItem(current_row)
    
    def setup_star_relation(self, layout):
        """设置星形关联的界面"""
        # 中心表选择
        center_layout = QHBoxLayout()
        center_layout.addWidget(QLabel("选择中心表:"))
        
        self.center_combo = QComboBox()
        for i, file_info in enumerate(self.file_data):
            self.center_combo.addItem(file_info['name'], i)
        center_layout.addWidget(self.center_combo, 1)
        
        layout.addLayout(center_layout)
        
        # 关联设置表格
        table = QTableWidget(0, 4)  # 初始为0行，实际行数会根据文件数量动态调整
        table.setHorizontalHeaderLabels(["关联表", "中心表字段", "关联表字段", "状态"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        layout.addWidget(table)
        self.star_table = table
        
        # 连接中心表选择变化事件
        self.center_combo.currentIndexChanged.connect(self.update_star_table)
        
        # 初始填充表格
        self.update_star_table(0)
    
    def update_star_table(self, index):
        """更新星形关联表格"""
        center_idx = self.center_combo.currentData()
        if center_idx is None or center_idx < 0 or center_idx >= len(self.file_data):
            return
            
        center_file = self.file_data[center_idx]
        
        # 计算需要多少行（除了中心表外的所有表）
        row_count = len(self.file_data) - 1
        self.star_table.setRowCount(row_count)
        
        # 填充表格
        row = 0
        for i, file_info in enumerate(self.file_data):
            if i == center_idx:
                continue  # 跳过中心表
                
            # 关联表名称
            name_item = QTableWidgetItem(file_info['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)  # 不可编辑
            self.star_table.setItem(row, 0, name_item)
            
            # 中心表字段下拉框
            center_combo = QComboBox()
            center_combo.addItems(center_file['columns'])
            center_combo.setProperty("row", row)
            center_combo.setProperty("related_idx", i)
            
            # 如果之前有设置，选择对应字段
            related_key = f'star_field_{i}'
            if related_key in center_file:
                idx = center_combo.findText(center_file[related_key])
                if idx >= 0:
                    center_combo.setCurrentIndex(idx)
            
            self.star_table.setCellWidget(row, 1, center_combo)
            
            # 关联表字段下拉框
            file_combo = QComboBox()
            file_combo.addItems(file_info['columns'])
            file_combo.setProperty("row", row)
            file_combo.setProperty("center_idx", center_idx)
            
            # 如果之前有设置，选择对应字段
            center_key = f'star_field_{center_idx}'
            if center_key in file_info:
                idx = file_combo.findText(file_info[center_key])
                if idx >= 0:
                    file_combo.setCurrentIndex(idx)
            
            self.star_table.setCellWidget(row, 2, file_combo)
            
            # 状态
            status_item = QTableWidgetItem("已设置")
            status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
            self.star_table.setItem(row, 3, status_item)
            
            row += 1
    
    def accept(self):
        """确认按钮处理"""
        if self.relation_type == "single":
            # 收集单一关联设置
            for i, combo in enumerate(self.single_field_combos):
                self.file_data[i]['ref_field'] = combo.currentText()
            
            self.result_data = {
                'relation_type': 'single',
                'file_data': self.file_data
            }
            
        elif self.relation_type == "chain":
            # 收集链式关联设置
            if not self.chain_relations:
                QMessageBox.warning(self, "警告", "请至少添加一个关联关系")
                return
                
            self.result_data = {
                'relation_type': 'chain',
                'chain_relations': self.chain_relations
            }
            
        else:  # star
            # 收集星形关联设置
            center_idx = self.center_combo.currentData()
            relationships = []
            
            for row in range(self.star_table.rowCount()):
                center_combo = self.star_table.cellWidget(row, 1)
                file_combo = self.star_table.cellWidget(row, 2)
                
                if center_combo and file_combo:
                    related_idx = center_combo.property("related_idx")
                    
                    relationships.append({
                        'center_idx': center_idx,
                        'related_idx': related_idx,
                        'center_field': center_combo.currentText(),
                        'related_field': file_combo.currentText()
                    })
            
            self.result_data = {
                'relation_type': 'star',
                'center_idx': center_idx,
                'relationships': relationships
            }
        
        super().accept()

class FormulaMergeHelper:
    """用于处理合并Excel文件时的公式调整"""
    
    @staticmethod
    def create_row_mapping(source_worksheet, target_worksheet, start_row_in_target):
        """创建源工作表到目标工作表的行映射关系"""
        row_mapping = {}
        # 从第2行开始（跳过标题行）
        for i, row in enumerate(source_worksheet.iter_rows(min_row=2), start=2):
            # 在目标工作表中的行号 = 起始行 + (源工作表中的行号 - 2)
            # -2是因为我们跳过了标题行，并且行号从1开始
            target_row = start_row_in_target + (i - 2)
            row_mapping[i] = target_row
        return row_mapping
    
    @staticmethod
    def parse_cell_references(formula):
        """解析公式中的单元格引用"""
        # 这是一个简化的实现，实际情况可能需要更复杂的解析
        references = []
        # 查找A1引用格式（例如A1, $A$1, A$1, $A1）
        # 注意：这种简单的正则表达式可能不足以处理所有复杂的Excel公式
        import re
        pattern = r'(\$?[A-Z]+\$?[0-9]+)'
        matches = re.findall(pattern, formula)
        for match in matches:
            references.append(match)
        return references
    
    @staticmethod
    def adjust_formula_references(formula, row_mapping):
        """调整公式中的行引用以匹配新的位置"""
        if not formula or not formula.startswith('='):
            return formula
            
        references = FormulaMergeHelper.parse_cell_references(formula)
        adjusted_formula = formula
        
        for ref in references:
            # 检查是否包含行号
            import re
            match = re.match(r'(\$?[A-Z]+)(\$?)([0-9]+)', ref)
            if match:
                col, dollar, row = match.groups()
                row_num = int(row)
                
                # 只有相对引用的行才需要调整
                if dollar != '$' and row_num in row_mapping:
                    new_row = row_mapping[row_num]
                    new_ref = f"{col}{dollar}{new_row}"
                    # 替换公式中的引用
                    adjusted_formula = adjusted_formula.replace(ref, new_ref)
        
        return adjusted_formula
    
    @staticmethod
    def copy_cell_with_adjusted_formula(source_cell, target_cell, row_mapping):
        """复制单元格的内容，如果是公式则调整行引用，避免StyleProxy错误"""
        # 完全避免复制样式，只复制值和公式
        # 处理值或公式
        if source_cell.value is not None:
            if isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                try:
                    # 是公式，调整并复制
                    adjusted_formula = FormulaMergeHelper.adjust_formula_references(source_cell.value, row_mapping)
                    target_cell.value = adjusted_formula
                except Exception as e:
                    print(f"调整公式引用时出错: {str(e)}")
                    # 出错时，直接复制原始值而不调整
                    target_cell.value = source_cell.value
            else:
                # 不是公式，直接复制值
                target_cell.value = source_cell.value
                
        # 尝试复制基本文本格式（不复制样式对象）
        try:
            # 尝试复制数字格式字符串
            if hasattr(source_cell, 'number_format') and source_cell.number_format:
                target_cell.number_format = source_cell.number_format
        except Exception as e:
            print(f"复制数字格式时出错: {str(e)}")

class ExcelMergerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.file_data = []  # 改为存储完整文件信息：[{'path':, 'name':, 'ref_col':, 'columns':}, ...]
        self.output_columns = []
        self.relation_data = None  # 存储关联设置
        self.ignoring_toggle = False  # 添加忽略切换事件的标志
        
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('Excel合并工具')
        self.setGeometry(100, 100, 1200, 800)
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        main_layout.setSpacing(5)  # 减少间距
        
        # 文件选择区域 - 进一步缩小
        file_group = QGroupBox("1. 选择要合并的Excel文件")
        file_layout = QHBoxLayout(file_group)
        file_layout.setContentsMargins(5, 5, 5, 5)  # 减少内边距
        
        self.btn_select_files = QPushButton('选择文件')
        self.btn_select_files.setFixedWidth(100)
        self.btn_select_files.clicked.connect(self.select_files)
        file_layout.addWidget(self.btn_select_files)
        
        self.file_list = QListWidget()
        self.file_list.setFixedHeight(40)  # 大幅减小高度
        file_layout.addWidget(self.file_list, 1)
        
        file_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        file_group.setMaximumHeight(80)  # 限制最大高度
        main_layout.addWidget(file_group)
        
        # 参考字段匹配区域 - 改为水平布局节省空间
        self.ref_group = QGroupBox("2. 设置参考字段")
        self.ref_group.setMaximumHeight(120)  # 限制最大高度
        ref_layout = QHBoxLayout(self.ref_group)  # 改为水平布局
        ref_layout.setContentsMargins(5, 5, 5, 5)  # 减少内边距
        
        # 使用固定小窗口放置表单
        ref_scroll = QScrollArea()
        ref_scroll.setWidgetResizable(True)
        ref_content = QWidget()
        self.ref_form = QFormLayout(ref_content)
        self.ref_form.setVerticalSpacing(2)  # 极小的垂直间距
        self.ref_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        ref_scroll.setWidget(ref_content)
        ref_layout.addWidget(ref_scroll)
        
        self.ref_group.setVisible(False)
        self.ref_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        main_layout.addWidget(self.ref_group)
        
        # 列选择区域 - 分配更多空间
        self.column_group = QGroupBox("3. 选择输出列")
        column_layout = QVBoxLayout(self.column_group)
        column_layout.setContentsMargins(5, 10, 5, 5)  # 适当的内边距
        
        # 创建全局操作区域 - 保持紧凑
        global_controls = QHBoxLayout()
        global_controls.setSpacing(5)
        
        search_label = QLabel("全局搜索:")
        global_controls.addWidget(search_label)
        
        self.column_search = QLineEdit()
        self.column_search.setPlaceholderText("在所有表格中搜索列名...")
        self.column_search.textChanged.connect(self.filter_columns)
        global_controls.addWidget(self.column_search, 1)
        
        global_select_all = QPushButton("全部全选")
        global_select_all.setFixedWidth(80)
        global_select_all.clicked.connect(self.select_all_global)
        global_controls.addWidget(global_select_all)
        
        global_deselect_all = QPushButton("全不选")
        global_deselect_all.setFixedWidth(80)
        global_deselect_all.clicked.connect(self.deselect_all_global)
        global_controls.addWidget(global_deselect_all)
        
        column_layout.addLayout(global_controls)
        
        # 主操作区 - 表格和已选列表水平排列
        main_selection = QHBoxLayout()
        
        # 左侧标签页 - 占据2/3空间
        self.column_tabs = QTabWidget()
        main_selection.addWidget(self.column_tabs, 2)
        
        # 右侧已选列表 - 占据1/3空间
        selected_box = QGroupBox("已选列(按选择顺序排列)")
        selected_layout = QVBoxLayout(selected_box)
        selected_layout.setContentsMargins(5, 10, 5, 5)
        
        self.selected_list = QListWidget()
        self.selected_list.setSelectionMode(QListWidget.ExtendedSelection)
        self.selected_list.setDragDropMode(QListWidget.InternalMove)  # 启用拖放排序
        selected_layout.addWidget(self.selected_list)
        
        # 添加上移下移按钮
        buttons_layout = QHBoxLayout()
        
        self.btn_move_up = QPushButton("上移")
        self.btn_move_up.clicked.connect(self.move_item_up)
        buttons_layout.addWidget(self.btn_move_up)
        
        self.btn_move_down = QPushButton("下移")
        self.btn_move_down.clicked.connect(self.move_item_down)
        buttons_layout.addWidget(self.btn_move_down)
        
        self.btn_clear = QPushButton("清空选择")
        self.btn_clear.clicked.connect(self.clear_selection)
        buttons_layout.addWidget(self.btn_clear)
        
        selected_layout.addLayout(buttons_layout)
        
        main_selection.addWidget(selected_box, 1)
        
        column_layout.addLayout(main_selection, 1)
        
        # 列选择组占据更多空间
        self.column_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.column_group.setVisible(False)
        main_layout.addWidget(self.column_group, 10)  # 分配10倍的拉伸比例
        
        # 输出设置和合并按钮放在一行 - 保持紧凑
        bottom_layout = QHBoxLayout()
        bottom_layout.setContentsMargins(0, 0, 0, 0)
        
        output_group = QGroupBox("4. 输出设置")
        output_layout = QVBoxLayout(output_group)  # 修改为垂直布局
        output_layout.setContentsMargins(5, 5, 5, 5)
        
        # 添加选项行
        options_layout = QHBoxLayout()
        
        # 修改复选框选项
        self.preserve_formulas_cb = QCheckBox("保留公式格式并自动应用到新数据")
        self.preserve_formulas_cb.setToolTip("勾选此选项将保留源文件中的Excel公式和格式，并在合并后应用到新增数据")
        self.preserve_formulas_cb.setChecked(True)  # 默认选中
        options_layout.addWidget(self.preserve_formulas_cb)
        
        output_layout.addLayout(options_layout)
        
        # 添加选择输出路径的行
        path_layout = QHBoxLayout()
        self.btn_output_path = QPushButton('选择输出路径')
        path_layout.addWidget(self.btn_output_path)
        self.btn_output_path.clicked.connect(self.select_output_path)
        
        self.output_path_label = QLabel('未选择输出路径')
        path_layout.addWidget(self.output_path_label, 1)
        
        output_layout.addLayout(path_layout)
        
        output_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        output_group.setMaximumHeight(100)  # 调整高度以适应新增的选项
        bottom_layout.addWidget(output_group, 2)
        
        self.btn_merge = QPushButton('开始合并')
        self.btn_merge.setMinimumWidth(120)
        self.btn_merge.setFixedHeight(40)
        self.btn_merge.clicked.connect(self.merge_files)
        self.btn_merge.setVisible(False)
        bottom_layout.addWidget(self.btn_merge)
        
        main_layout.addLayout(bottom_layout)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)
        
        # 添加状态标签
        self.status_label = QLabel("")
        main_layout.addWidget(self.status_label)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, '选择Excel文件', '', 'Excel文件 (*.xlsx *.xls)')
        
        if files:
            self.file_data = []  # 重置文件数据
            self.file_list.clear()
            
            # 读取每个文件的列
            for file_path in files:
                try:
                    df = pd.read_excel(file_path, nrows=1)
                    self.file_data.append({
                        'path': file_path,
                        'name': os.path.basename(file_path),
                        'columns': df.columns.tolist()
                    })
                    self.file_list.addItem(file_path)
                except Exception as e:
                    QMessageBox.critical(self, '错误', f'读取文件{file_path}失败: {str(e)}')
            
            if self.file_data:
                self.setup_relation_fields()
            
    def setup_relation_fields(self):
        """设置参考字段匹配界面"""
        # 清空原有内容
        while self.ref_form.rowCount() > 0:
            self.ref_form.removeRow(0)
        
        # 添加关联模式选择
        self.ref_form.addRow(QLabel("<b>选择关联模式:</b>"))
        mode_layout = QHBoxLayout()
        
        # 添加简单合并选项（排在最前面）
        self.simple_radio = QRadioButton("简单合并")
        self.simple_radio.clicked.connect(lambda: self.setup_simple_merge())
        mode_layout.addWidget(self.simple_radio)
        
        # 添加单一关联选项
        self.single_radio = QRadioButton("单一关联")
        # 使用clicked信号代替toggled
        self.single_radio.clicked.connect(lambda: self.open_relation_dialog("single"))
        mode_layout.addWidget(self.single_radio)
        
        # 链式关联
        self.chain_radio = QRadioButton("链式关联")
        # 使用clicked信号代替toggled
        self.chain_radio.clicked.connect(lambda: self.open_relation_dialog("chain"))
        mode_layout.addWidget(self.chain_radio)
        
        # 星形关联
        self.star_radio = QRadioButton("星形关联")
        # 使用clicked信号代替toggled
        self.star_radio.clicked.connect(lambda: self.open_relation_dialog("star"))
        mode_layout.addWidget(self.star_radio)
        
        # 添加重置按钮
        reset_btn = QPushButton("重置")
        reset_btn.setToolTip("重置关联模式和界面显示")
        reset_btn.clicked.connect(self.reset_relation_mode)
        mode_layout.addWidget(reset_btn)
        
        mode_widget = QWidget()
        mode_widget.setLayout(mode_layout)
        self.ref_form.addRow(mode_widget)
        
        # 显示参考字段区域
        self.ref_group.setVisible(True)
        self.ref_group.setMaximumHeight(120)  # 重置回默认高度
        self.btn_merge.setVisible(True)
        
        # 设置列选择界面
        self.setup_column_selection()
        
        # 重新显示列选择区域
        self.column_group.setVisible(True)
        self.column_group.setTitle("3. 选择输出列")  # 重置标题
        
    def reset_relation_mode(self):
        """重置关联模式和界面显示"""
        # 取消选中所有单选按钮
        self.simple_radio.setChecked(False)
        self.single_radio.setChecked(False)
        self.chain_radio.setChecked(False)
        self.star_radio.setChecked(False)
        
        # 清除关联数据
        self.relation_data = None
        
        # 清空第三部分的工作表选择标签页
        self.column_tabs.clear()
        
        # 重置第二部分的高度
        self.ref_group.setMaximumHeight(120)
        
        # 清除前面设置的工作表选择控件
        self.sheet_checkboxes = []
        
        # 重新设置列选择界面
        self.setup_column_selection()
        
        # 重新显示列选择区域，设置正确的标题
        self.column_group.setVisible(True)
        self.column_group.setTitle("3. 选择输出列")
        
        # 清空参考字段说明区域，只保留关联模式选择
        while self.ref_form.rowCount() > 2:  # 保留关联模式选择部分
            self.ref_form.removeRow(2)

    def setup_simple_merge(self):
        """设置简单合并模式"""
        # 清除之前的关联设置
        while self.ref_form.rowCount() > 2:  # 保留关联模式选择部分
            self.ref_form.removeRow(2)
        
        # 创建并保存简单合并的关联数据
        self.relation_data = {
            'relation_type': 'simple'
        }
        
        # 显示简单合并的说明
        self.ref_form.addRow(QLabel("<hr/><b>简单合并模式说明:</b>"))
        
        explanation = QLabel(
            "简单合并模式会查找所有Excel文件中名称相同的工作表(Sheet)，"
            "然后将它们的数据直接合并在一起。\n\n"
            "适用场景：多个Excel文件有相同结构的工作表，"
            "例如按部门拆分后的表格需要重新合并。\n\n"
            "注意：所有工作表的标题行(第一行)必须完全相同，合并时将自动去除重复的标题行。"
        )
        explanation.setWordWrap(True)
        self.ref_form.addRow(explanation)
        
        # 收集所有Excel文件中的所有工作表名称
        all_sheet_names = set()
        for file_info in self.file_data:
            try:
                excel = pd.ExcelFile(file_info['path'])
                all_sheet_names.update(excel.sheet_names)
            except Exception as e:
                print(f"读取文件 {file_info['name']} 的工作表列表失败: {str(e)}")
                
        if not all_sheet_names:
            error_label = QLabel("<font color='red'>未找到任何工作表！</font>")
            self.ref_form.addRow(error_label)
            return
        
        # 调整第二部分参考字段组的高度，为工作表选择提供更多空间
        self.ref_group.setMaximumHeight(400)  # 增加最大高度
        
        # 创建单独的工作表选择区域 - 第三部分
        self.column_group.setVisible(True)  # 显示第三部分
        self.column_group.setTitle("3. 选择输出列")  # 更新标题
        
        # 清空并重置第三部分
        for i in range(self.column_tabs.count()):
            self.column_tabs.removeTab(0)
        
        # 创建工作表选择标签页
        sheets_tab = QWidget()
        sheets_layout = QVBoxLayout(sheets_tab)
        
        # 添加操作按钮
        buttons_layout = QHBoxLayout()
        select_all_btn = QPushButton("全选")
        select_all_btn.setFixedWidth(80)
        buttons_layout.addWidget(select_all_btn)
        
        deselect_all_btn = QPushButton("全不选")
        deselect_all_btn.setFixedWidth(80)
        buttons_layout.addWidget(deselect_all_btn)
        
        # 添加搜索功能
        self.sheet_search = QLineEdit()
        self.sheet_search.setPlaceholderText("搜索工作表名称...")
        buttons_layout.addWidget(self.sheet_search, 1)
        
        sheets_layout.addLayout(buttons_layout)
        
        # 创建滚动区域容纳复选框
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setSpacing(5)
        
        # 添加工作表复选框
        self.sheet_checkboxes = []
        for sheet_name in sorted(all_sheet_names):
            cb = QCheckBox(sheet_name)
            cb.setChecked(True)  # 默认全选
            scroll_layout.addWidget(cb)
            self.sheet_checkboxes.append(cb)
        
        # 连接按钮事件
        select_all_btn.clicked.connect(lambda: self.toggle_all_sheets(True))
        deselect_all_btn.clicked.connect(lambda: self.toggle_all_sheets(False))
        self.sheet_search.textChanged.connect(self.filter_sheets)
        
        # 设置滚动区域
        scroll.setWidget(scroll_content)
        sheets_layout.addWidget(scroll)
        
        # 将工作表选择添加为一个标签页
        self.column_tabs.addTab(sheets_tab, "选择工作表")
        
        # 更新第三部分的标题
        self.column_group.setTitle("3. 选择需要合并的工作表")
        
        # 隐藏原来的已选列表部分
        for i in range(self.column_tabs.count()):
            widget = self.column_tabs.widget(i)
            if widget != sheets_tab:
                self.column_tabs.removeTab(i)
                break
        
        # 清空选择列表，但保留一个隐藏项以满足校验
        self.selected_list.clear()
        dummy_item = QListWidgetItem("__ALL_COLUMNS__")
        dummy_item.setData(Qt.UserRole, "__ALL_COLUMNS__")
        self.selected_list.addItem(dummy_item)
        
        # 显示合并按钮
        self.btn_merge.setVisible(True)
        
    def toggle_all_sheets(self, checked):
        """切换所有工作表的选中状态"""
        for cb in self.sheet_checkboxes:
            if cb.isVisible():  # 只操作可见的复选框
                cb.setChecked(checked)
    
    def filter_sheets(self, text):
        """过滤工作表显示"""
        for cb in self.sheet_checkboxes:
            cb.setVisible(text.lower() in cb.text().lower())

    def open_relation_dialog(self, relation_type):
        """打开关联设置对话框"""
        # 打开关联设置对话框
        dialog = RelationDialog(relation_type, self.file_data, self)
        if dialog.exec_() == QDialog.Accepted:
            # 保存关联设置
            self.relation_data = dialog.result_data
            
            # 更新界面提示
            self.update_relation_summary()
            
            # 显示合并按钮
            self.btn_merge.setVisible(True)
            
            # 显示列选择区域
            self.column_group.setVisible(True)
            self.column_group.setTitle("3. 选择输出列")  # 更新标题
        else:
            # 如果用户取消，取消选中当前按钮
            self.ignoring_toggle = True
            
            # 取消选中所有按钮
            self.simple_radio.setChecked(False)
            self.single_radio.setChecked(False)
            self.chain_radio.setChecked(False)
            self.star_radio.setChecked(False)
                
            # 重置标志
            self.ignoring_toggle = False
    
    def update_relation_summary(self):
        """更新关联设置摘要显示"""
        # 清空原有内容
        while self.ref_form.rowCount() > 2:  # 保留关联模式选择部分
            self.ref_form.removeRow(2)
        
        if not self.relation_data:
            return
        
        relation_type = self.relation_data.get('relation_type')
        
        # 添加摘要标题
        self.ref_form.addRow(QLabel("<hr/><b>已设置的关联关系:</b>"))
        
        if relation_type == 'simple':
            # 显示简单合并摘要
            summary = "简单合并模式：直接合并相同名称的工作表(Sheet)。\n"
            summary += "所有工作表的标题行将自动对齐，后续文件会跳过标题行。"
            
            summary_label = QLabel(summary)
            summary_label.setWordWrap(True)
            self.ref_form.addRow(summary_label)
            
        elif relation_type == 'single':
            # 显示单一关联摘要
            summary = "每个表格使用各自的关联字段："
            file_data = self.relation_data.get('file_data', [])
            
            for file_info in file_data:
                if 'ref_field' in file_info:
                    summary += f"\n• {file_info['name']}: {file_info['ref_field']}"
            
            summary_label = QLabel(summary)
            summary_label.setWordWrap(True)
            self.ref_form.addRow(summary_label)
            
        elif relation_type == 'chain':
            # 显示链式关联摘要
            relations = self.relation_data.get('chain_relations', [])
            
            if relations:
                for relation in relations:
                    source_idx = relation['source_idx']
                    target_idx = relation['target_idx']
                    source_field = relation['source_field']
                    target_field = relation['target_field']
                    
                    source_name = self.file_data[source_idx]['name']
                    target_name = self.file_data[target_idx]['name']
                    
                    text = f"{source_name}[{source_field}] → {target_name}[{target_field}]"
                    self.ref_form.addRow(QLabel(text))
            else:
                self.ref_form.addRow(QLabel("未设置任何关联关系"))
                
        elif relation_type == 'star':
            # 显示星形关联摘要
            center_idx = self.relation_data.get('center_idx')
            relationships = self.relation_data.get('relationships', [])
            
            if center_idx is not None and 0 <= center_idx < len(self.file_data):
                center_name = self.file_data[center_idx]['name']
                self.ref_form.addRow(QLabel(f"中心表: {center_name}"))
                
                for rel in relationships:
                    related_idx = rel['related_idx']
                    center_field = rel['center_field']
                    related_field = rel['related_field']
                    
                    if 0 <= related_idx < len(self.file_data):
                        related_name = self.file_data[related_idx]['name']
                        text = f"{center_name}[{center_field}] ⟷ {related_name}[{related_field}]"
                        self.ref_form.addRow(QLabel(text))
            else:
                self.ref_form.addRow(QLabel("未正确设置中心表和关联关系"))
        
        # 添加重新设置按钮
        reset_btn = QPushButton("重新设置关联关系")
        reset_btn.clicked.connect(self.setup_relation_fields)
        self.ref_form.addRow(reset_btn)

    def select_output_path(self):
        path, _ = QFileDialog.getSaveFileName(
            self, '保存合并文件', '', 'Excel文件 (*.xlsx)')
        
        if path:
            self.output_path = path
            self.output_path_label.setText(f'输出路径: {path}')
            
    def setup_column_selection(self):
        """使用标签页显示每个表格的列"""
        # 清空现有标签页
        self.column_tabs.clear()
        self.all_checkboxes = []  # 存储所有复选框以便全局操作
        self.selected_list.clear()  # 清空已选列表
        self.selection_order = []  # 新增: 跟踪选择顺序
        
        # 为每个表格创建一个标签页
        for idx, file_info in enumerate(self.file_data):
            file_name = file_info['name']
            columns = file_info['columns']
            
            # 创建标签页内容控件
            tab = QWidget()
            tab_layout = QVBoxLayout(tab)
            
            # 添加表格特定的搜索和全选控件
            controls_layout = QHBoxLayout()
            
            # 表格内搜索
            search = QLineEdit()
            search.setPlaceholderText(f"在 {file_name} 中搜索列...")
            controls_layout.addWidget(search, 1)
            
            # 表格内全选/全不选
            select_all = QPushButton("全选")
            deselect_all = QPushButton("全不选")
            controls_layout.addWidget(select_all)
            controls_layout.addWidget(deselect_all)
            
            tab_layout.addLayout(controls_layout)
            
            # 创建一个滚动区域容纳复选框
            scroll = QScrollArea()
            scroll.setWidgetResizable(True)
            
            # 内容容器
            content = QWidget()
            col_layout = QVBoxLayout(content)
            col_layout.setSpacing(5)  # 减小间距使更紧凑
            
            # 添加复选框
            checkboxes = []
            for col in columns:
                cb = QCheckBox(col)
                cb.setObjectName(f"{file_name}::{col}")  # 唯一标识
                # 连接勾选事件，更新选择顺序
                cb.stateChanged.connect(lambda state, c=cb: self.checkbox_changed(c, state))
                col_layout.addWidget(cb)
                checkboxes.append(cb)
                self.all_checkboxes.append(cb)  # 添加到全局列表
            
            # 连接事件
            select_all.clicked.connect(partial(self.select_all_checkboxes, checkboxes))
            deselect_all.clicked.connect(partial(self.deselect_all_checkboxes, checkboxes))
            search.textChanged.connect(partial(self.filter_checkboxes, checkboxes))
            
            # 设置滚动区域
            scroll.setWidget(content)
            tab_layout.addWidget(scroll)
            
            # 添加到标签页
            self.column_tabs.addTab(tab, f"{file_name}")
        
        # 显示列选择组
        self.column_group.setVisible(True)
    
    def filter_checkboxes(self, checkboxes, text):
        """过滤复选框显示"""
        for cb in checkboxes:
            cb.setVisible(text.lower() in cb.text().lower())
    
    def filter_columns(self, text):
        """全局搜索所有标签页中的列"""
        # 遍历所有标签页
        for i in range(self.column_tabs.count()):
            tab = self.column_tabs.widget(i)
            # 查找所有复选框并过滤
            for cb in tab.findChildren(QCheckBox):
                cb.setVisible(text.lower() in cb.text().lower())
        
        # 如果有搜索内容，自动选择包含匹配项最多的标签页
        if text:
            max_visible = 0
            max_tab = 0
            for i in range(self.column_tabs.count()):
                tab = self.column_tabs.widget(i)
                visible_count = sum(1 for cb in tab.findChildren(QCheckBox) if cb.isVisible())
                if visible_count > max_visible:
                    max_visible = visible_count
                    max_tab = i
            if max_visible > 0:
                self.column_tabs.setCurrentIndex(max_tab)

    def merge_files(self):
        if not hasattr(self, 'output_path'):
            QMessageBox.warning(self, '警告', '请先选择输出路径')
            return
            
        if not self.relation_data:
            QMessageBox.warning(self, '警告', '请先设置关联关系')
            return
            
        # 检查是否需要保留公式
        preserve_formulas = self.preserve_formulas_cb.isChecked()
        relation_type = self.relation_data.get('relation_type')
        
        # 开始合并
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("开始合并...")
        
        try:
            # 根据不同的关联模式和是否保留公式选择合并方法
            if relation_type == "simple":
                if preserve_formulas:
                    # 使用基于复制的方法保留格式和公式
                    self.merge_with_copy_and_append()
                else:
                    # 使用现有的pandas合并方法
                    self.merge_files_simple()
                return
            elif relation_type == "single":
                # 其他关联模式暂不支持保留公式
                if preserve_formulas:
                    QMessageBox.information(self, '提示', '目前仅支持简单合并模式下保留公式，将使用标准合并方法')
                merged_df = self.merge_files_single()
            elif relation_type == "chain":
                if preserve_formulas:
                    QMessageBox.information(self, '提示', '目前仅支持简单合并模式下保留公式，将使用标准合并方法')
                merged_df = self.merge_files_chain()
            else:
                if preserve_formulas:
                    QMessageBox.information(self, '提示', '目前仅支持简单合并模式下保留公式，将使用标准合并方法')
                merged_df = self.merge_files_star()
            
            # 以下是处理其他关联模式的现有代码
            if merged_df is None:
                self.progress_bar.setVisible(False)
                return
                
            # 获取所有选中的列（按照选择顺序）
            output_columns = []
            for i in range(self.selected_list.count()):
                item = self.selected_list.item(i)
                col_name = item.data(Qt.UserRole).split('::')[1]  # 获取原始列名
                output_columns.append(col_name)
            
            if not output_columns:
                QMessageBox.warning(self, '警告', '请至少选择一个输出列')
                return
                
            # 获取所有可能的列名
            available_columns = set(merged_df.columns)
            
            # 确保最终列顺序与用户选择顺序一致
            final_columns = []
            selected_output_cols = set()
            
            # 收集所有用户选择的列（包括重命名后的列）
            for col in output_columns:
                # 原始列名
                if col in available_columns:
                    final_columns.append(col)
                    selected_output_cols.add(col)
                else:
                    # 查找匹配的重命名列
                    for actual_col in available_columns:
                        if (actual_col.endswith(f"_{col}") or 
                            "_" in actual_col and actual_col.split("_", 1)[1] == col):
                            final_columns.append(actual_col)
                            selected_output_cols.add(actual_col)
                            break
            
            if len(final_columns) <= 0:
                QMessageBox.warning(self, '警告', '合并后没有找到选中的列')
                return
                
            # 创建最终数据框
            final_df = merged_df[final_columns]
            
            # 保存文件
            final_df.to_excel(self.output_path, index=False)
            
            QMessageBox.information(self, '完成', f'文件已成功合并并保存到: {self.output_path}')
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            QMessageBox.critical(self, '错误', f'合并失败: {str(e)}\n\n详细信息:\n{error_details}')
        finally:
            self.progress_bar.setVisible(False)

    def merge_files_single(self):
        """单一字段合并文件"""
        # 获取每个文件的关联字段
        file_data = self.relation_data.get('file_data', [])
        if not file_data:
            QMessageBox.critical(self, '错误', '未设置单一关联字段')
            return None
            
        merged_df = None
        
        for i, file_info in enumerate(file_data):
            self.progress_bar.setValue(i+1)
            QApplication.processEvents()
            
            file_path = file_info['path']
            file_name = file_info['name']
            ref_field = file_info.get('ref_field')
            
            if not ref_field:
                QMessageBox.critical(self, '错误', f'表 {file_name} 未设置关联字段')
                return None
            
            # 检查关联字段是否存在
            if ref_field not in file_info['columns']:
                QMessageBox.critical(self, '错误', f'表 {file_name} 中找不到关联字段: {ref_field}')
                return None
            
            # 读取当前文件
            df = pd.read_excel(file_path)
            
            # 如果是第一个文件，直接作为基础数据框
            if i == 0:
                merged_df = df
                continue
            
            # 获取上一个文件的关联字段
            prev_ref_field = file_data[0]['ref_field']  # 与第一个文件关联
            
            # 为每个表的所有列添加前缀（除关联字段外）
            rename_dict = {}
            for col in df.columns:
                if col == ref_field:
                    continue  # 保持关联字段不变
                elif col in merged_df.columns:
                    # 如果列名重复，添加文件名前缀
                    prefix = file_name.split('.')[0]  # 去掉扩展名
                    new_col_name = f"{prefix}_{col}"
                    rename_dict[col] = new_col_name
                else:
                    rename_dict[col] = col
            
            # 重命名列
            if rename_dict:
                df = df.rename(columns=rename_dict)
            
            # 执行合并
            merged_df = pd.merge(
                merged_df, df, 
                left_on=prev_ref_field, 
                right_on=ref_field, 
                how='outer'
            )
        
        return merged_df
    
    def merge_files_chain(self):
        """链式合并文件，根据用户自定义的关联关系"""
        # 获取链式关联设置
        relations = self.relation_data.get('chain_relations', [])
        if not relations:
            QMessageBox.critical(self, '错误', '未设置链式关联关系')
            return None
            
        # 创建一个图来表示表格间的关系
        graph = {}
        for relation in relations:
            source_idx = relation['source_idx']
            target_idx = relation['target_idx']
            
            if source_idx not in graph:
                graph[source_idx] = []
            graph[source_idx].append(relation)
        
        # 确定起始表格 (使用第一个关系的源表)
        start_idx = relations[0]['source_idx']
        
        # 读取起始表格
        start_file = self.file_data[start_idx]
        merged_df = pd.read_excel(start_file['path'])
        self.progress_bar.setValue(1)
        QApplication.processEvents()
            
        # 保存已处理的文件索引
        processed = {start_idx}
        
        # 按照关系逐步合并
        progress = 1
        while graph:
            # 查找下一个可合并的关系
            found = False
            for source_idx, relations_list in list(graph.items()):
                if source_idx in processed:
                    for relation in relations_list[:]:  # 使用副本遍历，以便能安全删除
                        target_idx = relation['target_idx']
                        
                        # 如果目标已处理，跳过
                        if target_idx in processed:
                            relations_list.remove(relation)
                            continue
                
                        # 找到一个可合并的关系
                        file_info = self.file_data[target_idx]
                        progress += 1
                        self.progress_bar.setValue(progress)
                        QApplication.processEvents()
                        
                        # 读取目标文件
                        df = pd.read_excel(file_info['path'])
                        
                        # 获取关联字段
                        source_field = relation['source_field']
                        target_field = relation['target_field']
            
                        # 为每个表的所有列添加前缀（除关联字段外）
                        rename_dict = {}
                        for col in df.columns:
                            if col == target_field:
                                continue  # 保持关联字段不变
                            elif col in merged_df.columns:
                                # 如果列名重复，添加文件名前缀
                                prefix = file_info['name'].split('.')[0]  # 去掉扩展名
                                new_col_name = f"{prefix}_{col}"
                                rename_dict[col] = new_col_name
                            else:
                                rename_dict[col] = col
            
                        # 重命名列
                        if rename_dict:
                            df = df.rename(columns=rename_dict)
            
                        # 执行合并
                        merged_df = pd.merge(
                            merged_df, df, 
                            left_on=source_field, 
                            right_on=target_field, 
                            how='outer'
                        )
                        
                        # 标记为已处理
                        processed.add(target_idx)
                        relations_list.remove(relation)
                        found = True
                        break
                        
                # 如果关系列表为空，从图中移除
                if not relations_list:
                    del graph[source_idx]
                    
                if found:
                    break
                    
            # 如果没有找到可合并的关系，说明可能有孤立的表格
            if not found and graph:
                # 尝试找到一个新的起点
                for source_idx in graph:
                    if source_idx not in processed:
                        # 读取该表格
                        file_info = self.file_data[source_idx]
                        df = pd.read_excel(file_info['path'])
                        
                        # 笛卡尔积合并
                        merged_df = pd.merge(
                            merged_df, df,
                            how='cross'  # 笛卡尔积
                        )
                        
                        processed.add(source_idx)
                        found = True
                        break
                        
                # 如果仍然没有找到，可能存在循环依赖，直接退出
                if not found:
                    break
        
        return merged_df
    
    def merge_files_star(self):
        """星形合并文件"""
        # 获取星形关联设置
        center_idx = self.relation_data.get('center_idx')
        relationships = self.relation_data.get('relationships', [])
        
        if center_idx is None or not relationships:
            QMessageBox.critical(self, '错误', '未设置星形关联关系')
            return None
        
        # 读取中心表
        center_info = self.file_data[center_idx]
        center_df = pd.read_excel(center_info['path'])
        self.progress_bar.setValue(1)
        QApplication.processEvents()
        
        # 逐个合并其他表
        for i, rel in enumerate(relationships):
            related_idx = rel['related_idx']
            center_field = rel['center_field']
            related_field = rel['related_field']
            
            file_info = self.file_data[related_idx]
            self.progress_bar.setValue(i+2)  # +2 因为中心表是第1个
            QApplication.processEvents()
            
            # 读取关联表
            df = pd.read_excel(file_info['path'])
            
            # 检查关联字段是否存在
            if center_field not in center_df.columns:
                QMessageBox.critical(self, '错误', f'中心表 {center_info["name"]} 中找不到关联字段: {center_field}')
                return None
                
            if related_field not in df.columns:
                QMessageBox.critical(self, '错误', f'表 {file_info["name"]} 中找不到关联字段: {related_field}')
                return None
            
            # 为每个表的所有列添加前缀（除关联字段外）
            rename_dict = {}
            for col in df.columns:
                if col == related_field:
                    continue  # 保持关联字段不变
                elif col in center_df.columns:
                    # 如果列名重复，添加文件名前缀
                    prefix = file_info['name'].split('.')[0]  # 去掉扩展名
                    new_col_name = f"{prefix}_{col}"
                    rename_dict[col] = new_col_name
                else:
                    rename_dict[col] = col
            
            # 重命名列
            if rename_dict:
                df = df.rename(columns=rename_dict)
            
            # 执行合并
            center_df = pd.merge(
                center_df, df, 
                left_on=center_field, 
                right_on=related_field, 
                how='outer'
            )
        
        return center_df

    def merge_files_simple(self):
        """简单合并模式 - 合并相同名称的工作表"""
        # 显示进度条
        QApplication.processEvents()
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(5)
        
        try:
            # 获取所有被选中的工作表
            selected_sheets = []
            for cb in self.sheet_checkboxes:
                if cb.isChecked():
                    selected_sheets.append(cb.text())
            
            if not selected_sheets:
                QMessageBox.warning(self, '警告', '请至少选择一个工作表进行合并!')
                self.progress_bar.setVisible(False)
                return
            
            # 显示已选择的工作表数量
            print(f"已选择 {len(selected_sheets)} 个工作表进行合并")
            
            # 第一步：为每个选中的工作表名称创建一个字典，记录包含该工作表的所有文件
            sheet_files = {sheet: [] for sheet in selected_sheets}
            
            for i, file_info in enumerate(self.file_data):
                self.progress_bar.setValue(5 + (i * 15 // len(self.file_data)))
                QApplication.processEvents()
                
                file_path = file_info['path']
                file_name = file_info['name']
                
                try:
                    # 使用pandas读取所有工作表名称
                    excel = pd.ExcelFile(file_path)
                    file_sheets = set(excel.sheet_names)
                    
                    # 检查该文件中包含哪些选中的工作表
                    for sheet in selected_sheets:
                        if sheet in file_sheets:
                            sheet_files[sheet].append(file_path)
                            print(f"文件 {file_name} 包含工作表 {sheet}")
                    
                except Exception as e:
                    QMessageBox.warning(self, '警告', f'读取文件 {file_name} 的工作表列表失败: {str(e)}')
            
            # 第二步：对每个选中的工作表进行合并处理
            self.progress_bar.setValue(20)
            QApplication.processEvents()
            
            total_sheets = len(selected_sheets)
            merged_results = {}
            
            with pd.ExcelWriter(self.output_path) as writer:
                for idx, sheet_name in enumerate(selected_sheets):
                    # 更新进度
                    progress = 20 + (idx * 80 // total_sheets)
                    self.progress_bar.setValue(progress)
                    QApplication.processEvents()
                    
                    # 获取包含该工作表的所有文件
                    files = sheet_files[sheet_name]
                    print(f"正在合并工作表: {sheet_name}, 在 {len(files)} 个文件中存在")
                    
                    if not files:
                        print(f"工作表 {sheet_name} 没有有效文件，跳过")
                        continue
                    
                    try:
                        # 创建一个列表存储所有文件的数据
                        all_dfs = []
                        
                        # 读取每个文件中的这个工作表
                        for file_idx, file_path in enumerate(files):
                            try:
                                # 直接读取整个工作表，不区分标题行和数据行
                                df = pd.read_excel(file_path, sheet_name=sheet_name)
                                
                                # 如果工作表为空，跳过
                                if df.empty:
                                    print(f"文件 {file_path} 的工作表 {sheet_name} 为空，跳过")
                                    continue
                                    
                                print(f"文件 {file_path} 的工作表 {sheet_name} 包含 {len(df)} 行数据")
                                all_dfs.append(df)
                                
                            except Exception as e:
                                print(f"处理文件 {file_path} 的工作表 {sheet_name} 时出错: {str(e)}")
                                continue
                        
                        if not all_dfs:
                            print(f"工作表 {sheet_name} 没有有效数据，跳过")
                            continue
                            
                        # 合并所有数据框
                        print(f"合并 {len(all_dfs)} 个文件的 {sheet_name} 工作表数据")
                        merged_df = pd.concat(all_dfs, ignore_index=True)
                        print(f"合并后共有 {len(merged_df)} 行数据")
                        
                        # 删除重复行
                        original_count = len(merged_df)
                        merged_df = merged_df.drop_duplicates()
                        print(f"删除了 {original_count - len(merged_df)} 行重复数据")
                        
                        # 如果成功合并了数据，写入到输出文件
                        if not merged_df.empty:
                            merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            merged_results[sheet_name] = len(merged_df)
                            print(f"成功合并工作表 {sheet_name}, 共 {len(merged_df)} 行数据")
                        else:
                            print(f"工作表 {sheet_name} 合并后没有数据")
                    
                    except Exception as e:
                        error_msg = f"合并工作表 {sheet_name} 时出错: {str(e)}"
                        print(error_msg)
                        QMessageBox.warning(self, '警告', error_msg)
            
            # 显示成功信息
            self.progress_bar.setValue(100)
            QApplication.processEvents()
            
            if merged_results:
                success_msg = f"成功合并 {len(merged_results)} 个工作表到文件: {self.output_path}\n\n"
                success_msg += "合并详情:\n"
                
                for sheet, rows in merged_results.items():
                    success_msg += f"- {sheet}: {rows} 行数据\n"
                
                QMessageBox.information(self, '完成', success_msg)
            else:
                QMessageBox.warning(self, '警告', '未能成功合并任何工作表')
        
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            error_msg = f'合并过程中发生错误: {str(e)}\n\n详细信息:\n{error_details}'
            print(error_msg)
            QMessageBox.critical(self, '错误', error_msg)
        
        finally:
            # 隐藏进度条
            self.progress_bar.setVisible(False)

    def merge_with_copy_and_append(self):
        """基于复制第一个文件并追加数据的合并方法"""
        # 更新进度条
        self.progress_bar.setValue(5)
        QApplication.processEvents()
        
        try:
            # 获取所有被选中的工作表
            selected_sheets = []
            for cb in self.sheet_checkboxes:
                if cb.isChecked():
                    selected_sheets.append(cb.text())
            
            if not selected_sheets:
                QMessageBox.warning(self, '警告', '请至少选择一个工作表进行合并!')
                self.progress_bar.setVisible(False)
                return
            
            # 获取第一个文件的路径
            first_file = self.file_data[0]['path']
            
            # 复制第一个文件到输出路径
            self.status_label.setText("正在复制基础文件...")
            shutil.copy2(first_file, self.output_path)
            
            # 更新进度条
            self.progress_bar.setValue(20)
            QApplication.processEvents()
            
            # 使用openpyxl打开复制的文件
            target_workbook = openpyxl.load_workbook(self.output_path)
            
            # 总工作表计数以更新进度
            total_sheets = len(selected_sheets)
            sheets_progress_portion = 70  # 工作表处理占总进度的70%
            
            # 记录处理结果
            merge_results = {}
            
            # 逐个处理选中的工作表
            for idx, sheet_name in enumerate(selected_sheets):
                # 更新进度和状态
                progress = 20 + (idx * sheets_progress_portion // total_sheets)
                self.progress_bar.setValue(progress)
                self.status_label.setText(f"正在处理工作表: {sheet_name}...")
                QApplication.processEvents()
                
                # 检查目标工作簿中是否存在该工作表
                if sheet_name not in target_workbook.sheetnames:
                    print(f"目标文件中不存在工作表 {sheet_name}, 跳过")
                    continue
                
                target_sheet = target_workbook[sheet_name]
                
                # 找到这个工作表在目标文件中的最后一行
                max_row = target_sheet.max_row
                
                # 初始化结果统计
                if sheet_name not in merge_results:
                    merge_results[sheet_name] = {
                        'original_rows': max_row,
                        'added_rows': 0
                    }
                
                # 缓存第一个文件的格式和公式信息（用于后续应用）
                first_row_format = {}
                
                # 只检查前10行来寻找公式和格式
                formula_rows = []
                for r in range(1, min(max_row + 1, 11)):
                    has_formula = False
                    for c in range(1, target_sheet.max_column + 1):
                        cell = target_sheet.cell(row=r, column=c)
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            has_formula = True
                            break
                    if has_formula:
                        formula_rows.append(r)
                        
                # 保存最后一个包含数据的行的格式（用于复制格式）
                if max_row > 0:
                    for c in range(1, target_sheet.max_column + 1):
                        first_row_format[c] = target_sheet.cell(row=max_row, column=c)
                
                # 从第二个文件开始处理
                for file_idx in range(1, len(self.file_data)):
                    file_info = self.file_data[file_idx]
                    file_path = file_info['path']
                    file_name = file_info['name']
                    
                    # 更新子进度
                    sub_progress = progress + ((file_idx) * sheets_progress_portion // (len(self.file_data) * total_sheets))
                    self.progress_bar.setValue(sub_progress)
                    self.status_label.setText(f"处理工作表 {sheet_name}, 文件 {file_name}...")
                    QApplication.processEvents()
                    
                    try:
                        # 检查源文件是否包含该工作表
                        excel = pd.ExcelFile(file_path)
                        if sheet_name not in excel.sheet_names:
                            print(f"文件 {file_name} 中不存在工作表 {sheet_name}, 跳过")
                            continue
                        
                        # 读取源工作表数据
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        
                        # 如果工作表为空，跳过
                        if df.empty:
                            print(f"文件 {file_name} 的工作表 {sheet_name} 为空，跳过")
                            continue
                        
                        # 不跳过第一行，保留所有数据行
                        data_rows = df.values.tolist()
                        
                        # 将数据行追加到目标工作表
                        start_row = max_row + 1
                        for row_idx, row_data in enumerate(data_rows):
                            current_row = start_row + row_idx
                            for col_idx, value in enumerate(row_data, 1):
                                cell = target_sheet.cell(row=current_row, column=col_idx)
                                cell.value = value
                                
                                # 应用格式
                                if col_idx in first_row_format:
                                    self.copy_cell_format(first_row_format[col_idx], cell)
                        
                        # 应用公式
                        if formula_rows:
                            rows_added = len(data_rows)
                            self.status_label.setText(f"应用公式到新添加的 {rows_added} 行数据...")
                            QApplication.processEvents()
                            
                            for formula_row in formula_rows:
                                for col_idx in range(1, target_sheet.max_column + 1):
                                    source_cell = target_sheet.cell(row=formula_row, column=col_idx)
                                    
                                    # 检查是否是公式
                                    if source_cell.value and isinstance(source_cell.value, str) and source_cell.value.startswith('='):
                                        # 为新增的每一行应用公式
                                        for r in range(start_row, start_row + rows_added):
                                            # 计算行偏移量
                                            row_diff = r - formula_row
                                            # 复制公式并调整行引用
                                            target_cell = target_sheet.cell(row=r, column=col_idx)
                                            adjusted_formula = self.adjust_formula_for_row(source_cell.value, row_diff)
                                            target_cell.value = adjusted_formula
                                            # 复制格式
                                            self.copy_cell_format(source_cell, target_cell)
                        
                        # 更新最后一行位置
                        max_row += len(data_rows)
                        
                        # 更新统计信息
                        merge_results[sheet_name]['added_rows'] += len(data_rows)
                        
                        print(f"从文件 {file_name} 添加了 {len(data_rows)} 行数据到工作表 {sheet_name}")
                        
                    except Exception as e:
                        error_msg = f"处理文件 {file_name} 的工作表 {sheet_name} 时出错: {str(e)}"
                        print(error_msg)
                        # 继续处理其他文件
            
            # 保存更改
            self.progress_bar.setValue(95)
            self.status_label.setText("正在保存合并结果...")
            QApplication.processEvents()
            target_workbook.save(self.output_path)
            
            # 显示成功信息
            self.progress_bar.setValue(100)
            self.status_label.setText("合并完成！")
            QApplication.processEvents()
            
            # 准备成功消息
            success_msg = f"成功合并数据到文件: {self.output_path}\n\n"
            success_msg += "合并详情:\n"
            
            for sheet, result in merge_results.items():
                original = result['original_rows']
                added = result['added_rows']
                success_msg += f"- {sheet}: 原始 {original} 行, 新增 {added} 行, 总计 {original + added} 行\n"
            
            if self.preserve_formulas_cb.isChecked():
                success_msg += "\n已自动应用格式和填充公式到新增数据。"
            
            QMessageBox.information(self, '合并完成', success_msg)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            error_msg = f'合并过程中发生错误: {str(e)}\n\n详细信息:\n{error_details}'
            print(error_msg)
            QMessageBox.critical(self, '错误', error_msg)
        
        finally:
            # 恢复界面状态
            self.btn_merge.setEnabled(True)
            self.progress_bar.setVisible(False)
    
    def copy_cell_format(self, source_cell, target_cell):
        """从源单元格复制格式到目标单元格"""
        try:
            if source_cell.has_style:
                # 复制字体
                target_cell.font = copy.copy(source_cell.font)
                # 复制对齐方式
                target_cell.alignment = copy.copy(source_cell.alignment)
                # 复制边框
                target_cell.border = copy.copy(source_cell.border)
                # 复制填充
                target_cell.fill = copy.copy(source_cell.fill)
                # 复制数字格式
                target_cell.number_format = source_cell.number_format
        except Exception as e:
            print(f"复制单元格格式时出错: {str(e)}")
    
    def adjust_formula_for_row(self, formula, row_diff):
        """调整公式以适应新行，修改行引用"""
        if not formula or not formula.startswith('='):
            return formula
            
        # 解析并调整单元格引用
        import re
        adjusted_formula = formula
        
        # 匹配A1类型的单元格引用
        cell_pattern = r'([A-Z]+)(\d+)'
        matches = re.findall(cell_pattern, formula)
        
        for match in matches:
            col_ref, row_ref = match
            old_ref = f"{col_ref}{row_ref}"
            row_num = int(row_ref)
            
            # 调整行号
            new_row_num = row_num + row_diff
            new_ref = f"{col_ref}{new_row_num}"
            
            # 替换公式中的引用
            adjusted_formula = adjusted_formula.replace(old_ref, new_ref)
        
        return adjusted_formula

    # 添加新的辅助函数，修复lambda问题
    def select_all_checkboxes(self, checkboxes):
        for cb in checkboxes:
            if cb.isVisible():
                cb.setChecked(True)

    def deselect_all_checkboxes(self, checkboxes):
        for cb in checkboxes:
            cb.setChecked(False)

    # 添加全局全选/全不选方法
    def select_all_global(self):
        for cb in self.all_checkboxes:
            if cb.isVisible():
                cb.setChecked(True)

    def deselect_all_global(self):
        for cb in self.all_checkboxes:
            if cb.isVisible():
                cb.setChecked(False)

    def clear_selection(self):
        """清空所有选择"""
        # 取消所有复选框的选中状态
        for cb in self.all_checkboxes:
            cb.setChecked(False)
        
        # 清空已选列表
        self.selected_list.clear()

    def checkbox_changed(self, checkbox, state):
        """当复选框状态改变时更新已选列表，保持选择顺序"""
        column_text = checkbox.text()
        source_file = checkbox.objectName().split('::')[0]  # 获取来源文件
        
        if state == Qt.Checked:
            # 添加到已选列表和选择顺序
            item = QListWidgetItem(f"{column_text} (来自: {source_file})")
            item.setData(Qt.UserRole, checkbox.objectName())  # 存储唯一标识
            self.selected_list.addItem(item)
        else:
            # 从已选列表中删除
            for i in range(self.selected_list.count()):
                item = self.selected_list.item(i)
                if item.data(Qt.UserRole) == checkbox.objectName():
                    self.selected_list.takeItem(i)
                    break

    def move_item_up(self):
        """将选中项向上移动"""
        current_row = self.selected_list.currentRow()
        if current_row > 0:
            item = self.selected_list.takeItem(current_row)
            self.selected_list.insertItem(current_row - 1, item)
            self.selected_list.setCurrentItem(item)

    def move_item_down(self):
        """将选中项向下移动"""
        current_row = self.selected_list.currentRow()
        if current_row < self.selected_list.count() - 1:
            item = self.selected_list.takeItem(current_row)
            self.selected_list.insertItem(current_row + 1, item)
            self.selected_list.setCurrentItem(item)

def main():
    app = QApplication(sys.argv)
    ex = ExcelMergerApp()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    # 是否为演示模式
    DEMO_MODE = False
    main()
